from athlete_data import AthleteData
from athlete_data import timeStrToSeconds
from athlete_data import writeAtheletesToCSV
from athlete_data import fillExcelWorksheet
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select
from openpyxl import Workbook, load_workbook
from copy import copy
import requests
import time
import os


# Returns the index of the option with the given string
def findOptionIndex(elem, name: str):
    option_index = -1
    options = elem.find_elements(By.TAG_NAME, "option")
    numOptions = len(options)
    for index in range(0, numOptions):
        option = options[index]
        if name in option.text:
            option_index = index
            break
    
    return option_index


# This function scrapes the info of a single athlete
def ScrapeAthleteInfo(url: str) -> AthleteData:
    athlete = None
    try:
        # get the HTML to scrape
        response = requests.get(url)
        html = response.text

        # use BeautifulSoup
        soup = BeautifulSoup(html, 'lxml')

        # first find the athlete data
        allTitles = soup.find_all('h3')
        participant_title = None
        members_title = None
        scoring_title = None
        doublesTeam_title = None
        workout_title = None
        judging_title = None
        overall_time_title = None
        for titleH3 in allTitles:
            if ("Participant" in titleH3.text):
                participant_title = titleH3
            if ("Members" in titleH3.text):
                members_title = titleH3
            elif ("Scoring" in titleH3.text):
                scoring_title = titleH3
            elif ("Team" in titleH3.text):
                doublesTeam_title = titleH3
            elif ("Workout" in titleH3.text):
                workout_title = titleH3
            elif ("Judging" in titleH3.text):
                judging_title = titleH3
            elif ("Overall Time" in titleH3.text):
                overall_time_title = titleH3
            
        # gather info on the participant
        participantName = None
        ageGroup = ""
        eventName = ""
        division = ""
        if participant_title != None:
            participantRoot = participant_title.parent
            participantTable = participantRoot.table.tbody
            for row in participantTable.find_all('tr'):
                if row.th != None:
                    if "Name" in row.th.text:
                        participantName = row.td.text
                    elif "Age Group" in row.th.text:
                        ageGroup = row.td.text

        if members_title != None:
            membersRoot = members_title.parent
            membersTable = membersRoot.table.tbody
            for row in membersTable.find_all('tr'):
                if row.th != None:
                    if "Member" in row.th.text:
                        if (participantName == None):
                            participantName = row.td.text
                        else:
                            participantName = participantName + " - " + row.td.text

        if doublesTeam_title != None:
            doublesTeamRoot = doublesTeam_title.parent
            doublesTeamTable = doublesTeamRoot.table.tbody
            for row in doublesTeamTable.find_all('tr'):
                if row.th != None:
                    if "Age Group" in row.th.text:
                        ageGroup = row.td.text
                    elif "Division" in row.th.text:
                        division = row.td.text

        if scoring_title != None:
            # extra info (Race and Division)
            scoringRoot = scoring_title.parent
            scoringTable = scoringRoot.table.tbody
            for row in scoringTable.find_all('tr'):
                if row.th != None:
                    if "Race" in row.th.text:
                        eventName = row.td.text
                    elif "Division" in row.th.text:
                        division = row.td.text


        athlete = None

        if participantName != None and workout_title != None and judging_title != None and overall_time_title != None:
            print("Scraping info for ", participantName , " (", ageGroup, ")")

            athlete = AthleteData()
            athlete.name = participantName
            athlete.ageGroup = ageGroup
            athlete.event = eventName
            athlete.division = division

            # Judging
            judgingRoot = judging_title.parent
            judgingTable = judgingRoot.table.tbody
            for row in judgingTable.find_all('tr'):
                if row.th != None:
                    if "Penalty" in row.th.text:
                        athlete.penalty = row.td.text

            # Overall Time
            overallRoot = overall_time_title.parent
            overallTable = overallRoot.table.tbody
            for row in overallTable.find_all('tr'):
                if row.th != None:
                    if "Overall Time" in row.th.text:
                        timeStr = row.td.text
                        athlete.totalTime = timeStrToSeconds(timeStr)

            # the workout results
            workoutRoot = workout_title.parent
            workoutTable = workoutRoot.table.tbody
            for row in workoutTable.find_all('tr'):
                if row.th != None:
                    rowName = row.th.text
                    if "Running" in rowName:
                        # first get the index of the running lap
                        runningIndexChar = rowName[-1]
                        runningIndex = ord(runningIndexChar) - ord('1')

                        # get the time
                        runningTimeStr = row.td.text
                        runningTime = timeStrToSeconds(runningTimeStr)
                        athlete.running[runningIndex] = runningTime
                    elif "SkiErg" in rowName:
                        timeStr = row.td.text
                        athlete.skierg = timeStrToSeconds(timeStr)
                    elif "Sled Push" in rowName:
                        timeStr = row.td.text
                        athlete.sledPush = timeStrToSeconds(timeStr)
                    elif "Sled Pull" in rowName:
                        timeStr = row.td.text
                        athlete.sledPull = timeStrToSeconds(timeStr)
                    elif "Burpee" in rowName:
                        timeStr = row.td.text
                        athlete.burpeeBroadJump = timeStrToSeconds(timeStr)
                    elif "Row" in rowName:
                        timeStr = row.td.text
                        athlete.row = timeStrToSeconds(timeStr)
                    elif "Farmer" in rowName:
                        timeStr = row.td.text
                        athlete.farmersCarry = timeStrToSeconds(timeStr)
                    elif "Lunges" in rowName:
                        timeStr = row.td.text
                        athlete.sandbagLunges = timeStrToSeconds(timeStr)
                    elif "Ball" in rowName:
                        timeStr = row.td.text
                        athlete.wallBalls = timeStrToSeconds(timeStr)
                    elif "Roxzone" in rowName:
                        timeStr = row.td.text
                        athlete.roxzoneTime = timeStrToSeconds(timeStr)
                    elif "Run Total" in rowName:
                        timeStr = row.td.text
                        athlete.runTotal = timeStrToSeconds(timeStr)
            
            #print(athlete.getCSVLine())

    except Exception as e:
        print('Error scraping info for athlete. URL = ', url)
        print(e)

    return athlete

# This function scrapes an event at the given division for the given sex
def ScrapeHyroxResults(driver, eventName: str, Division: str, Sex: str) -> list:

    print("Scraping Hyrox results for ", eventName, " - ", Division, " - ", Sex)

    division_index = 0
    match Division:
        case "HYROX PRO":
            division_index = 2
        case "HYROX":
            division_index = 5
        case "HYROX PRO DOUBLES":
            division_index = 8
        case "HYROX DOUBLES":
            division_index = 11
        case "HYROX TEAM RELAY":
            division_index = 14
    
    driver.get('https://results.hyrox.com/season-7/?pid=start&pidp=ranking_nav')
    time.sleep(0.25)
    race_selector_elem = driver.find_element(by=By.ID, value="default-lists-event_main_group")
    race_selector = Select(race_selector_elem)
    option_index = findOptionIndex(race_selector_elem, eventName)
    if option_index >= 0:
        race_selector.select_by_index(option_index)
    time.sleep(0.25)

    division_selector = Select(driver.find_element(by=By.ID, value="default-lists-event"))
    division_selector.select_by_index(division_index)    # 0 is HYROX PRO, 1 is HYROX TEAM RELAY, 2 is HYROX, 3 is HYROX DOUBLES
    time.sleep(0.25)

    gender_selector_elem = driver.find_element(by=By.ID, value="default-lists-sex")
    gender_selector = Select(gender_selector_elem)
    option_index = findOptionIndex(gender_selector_elem, Sex)
    if option_index >= 0:
        gender_selector.select_by_index(option_index)
    time.sleep(0.25)

    results_per_page_selector = Select(driver.find_element(by=By.ID, value="default-num_results"))
    results_per_page_selector.select_by_index(2)         # 0 is 25, 1 is 50, 2 is 100
    time.sleep(0.25)

    search_button = driver.find_element(by=By.ID, value="default-submit")
    search_button.click()

    time.sleep(1)

    # now we should iterate over all the links to athletes
    athletes = []

    nextPageLink = ""
    while (nextPageLink != None):
        mainContentElem = driver.find_element(by=By.ID, value="cbox-main")
        allNameElements = mainContentElem.find_elements(by=By.TAG_NAME, value="a")
        nextPageLink = None
        for athleteLinkElem in allNameElements:
            try:
                linkPage = athleteLinkElem.get_attribute('href')
                if (not 'favorite_add' in linkPage and
                    '&idp=' in linkPage):
                    #print("Analyzing athlete: ", athleteLinkElem.text)
                    newAthlete = ScrapeAthleteInfo(linkPage)
                    newAthlete.event = eventName        # ensure the event name is correct
                    if (newAthlete.name != None):
                        athletes.append(newAthlete)
                elif ('?page=' in linkPage and
                    athleteLinkElem.text == '>'):
                    nextPageLink = athleteLinkElem

            except:
                pass
            
        # check if there's another page to analyze
        if (nextPageLink != None):
            nextPageLink.click()
            time.sleep(1)

    # return the list of athletes
    return athletes


# Copies the cells from source_sheet to target_sheet
def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

    target_sheet.conditional_formatting = source_sheet.conditional_formatting
        

def scrapeHyroxCompleteEvent(driver, eventName: str, excelFilePath: str):
    # Create an Excel file
    workbook = Workbook()
    worksheet = workbook.active

    # Copy the stats sheet from the reference sheet
    reference_workbook = load_workbook('data\\RefHyroxSheet.xlsx')
    reference_sheet = reference_workbook.get_sheet_by_name("Stats")
    worksheet.title = "Stats"
    copy_cells(reference_sheet, worksheet)

    # HYROX PRO Men
    athletes = ScrapeHyroxResults(driver, eventName, 'HYROX PRO', 'Men')
    worksheet = workbook.create_sheet("HYROX_PRO_Men")
    fillExcelWorksheet(worksheet, athletes)

    # HYROX PRO Women
    athletes = ScrapeHyroxResults(driver, eventName, 'HYROX PRO', 'Women')
    worksheet = workbook.create_sheet("HYROX_PRO_Women")
    fillExcelWorksheet(worksheet, athletes)

    # HYROX Men
    athletes = ScrapeHyroxResults(driver, eventName, 'HYROX', 'Men')
    worksheet = workbook.create_sheet("HYROX_Men")
    fillExcelWorksheet(worksheet, athletes)

    # HYROX Women
    athletes = ScrapeHyroxResults(driver, eventName, 'HYROX', 'Women')
    worksheet = workbook.create_sheet("HYROX_Women")
    fillExcelWorksheet(worksheet, athletes)

    # HYROX PRO DOUBLES Men
    athletes = ScrapeHyroxResults(driver, eventName, 'HYROX PRO DOUBLES', 'Men')
    worksheet = workbook.create_sheet("HYROX_PRO_DOUBLES_Men")
    fillExcelWorksheet(worksheet, athletes)

    # HYROX PRO DOUBLES Women
    athletes = ScrapeHyroxResults(driver, eventName, 'HYROX PRO DOUBLES', 'Women')
    worksheet = workbook.create_sheet("HYROX_PRO_DOUBLES_Women")
    fillExcelWorksheet(worksheet, athletes)

    # HYROX DOUBLES Men
    athletes = ScrapeHyroxResults(driver, eventName, 'HYROX DOUBLES', 'Men')
    worksheet = workbook.create_sheet("HYROX_DOUBLES_Men")
    fillExcelWorksheet(worksheet, athletes)

    # HYROX DOUBLES Women
    athletes = ScrapeHyroxResults(driver, eventName, 'HYROX DOUBLES', 'Women')
    worksheet = workbook.create_sheet("HYROX_DOUBLES_Women")
    fillExcelWorksheet(worksheet, athletes)

    # HYROX DOUBLES Mixed
    athletes = ScrapeHyroxResults(driver, eventName, 'HYROX DOUBLES', 'Mixed')
    worksheet = workbook.create_sheet("HYROX_DOUBLES_Mixed")
    fillExcelWorksheet(worksheet, athletes)

    # HYROX TEAM RELAY
    athletes = ScrapeHyroxResults(driver, eventName, 'HYROX TEAM RELAY', 'Men')
    worksheet = workbook.create_sheet("HYROX_TEAM_RELAY")
    fillExcelWorksheet(worksheet, athletes)

    workbook.save(excelFilePath)


# create the Selenium web driver
driver = webdriver.Chrome()

# scrape the whole Valencia 2023 event
currFolder = os.getcwd()
filePath = currFolder + '\\data\\2024-HyroxValencia.xlsx'
scrapeHyroxCompleteEvent(driver, '2024 Valencia', filePath)

# close the browser
driver.quit()
