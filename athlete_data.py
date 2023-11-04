import re
import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from datetime import timedelta

# Converts a string with format '##:##:##' to the total seconds
def timeStrToSeconds(str: str) -> int:
    if (len(str) <= 1):
        return 0
    
    parts = re.findall(r'\d+', str)
    hours = int(parts[0])
    minutes = int(parts[1])
    seconds = int(parts[2])
    runningTime = hours * 3600 + minutes * 60 + seconds

    return runningTime

# Converts an integer to a time with format '##:##:##' considering the time as seconds
def timeSecondsToStr(totalSeconds: int) -> str:
    seconds = totalSeconds
    hours = int(seconds // 3600)
    seconds -= hours * 3600
    minutes = int(seconds // 60)
    seconds -= minutes * 60
    seconds = int(seconds)
    timeStr = str(hours).zfill(2) + ":" + str(minutes).zfill(2) + ":" + str(seconds).zfill(2)

    return timeStr

# Returns the deltatime associated to the given total seconds
def timeSecondsToDeltaTime(totalSeconds: int):
    seconds = totalSeconds
    hours = int(seconds // 3600)
    seconds -= hours * 3600
    minutes = int(seconds // 60)
    seconds -= minutes * 60
    seconds = int(seconds)
    timeDelta = timedelta(hours=hours, minutes=minutes, seconds=seconds)

    return timeDelta


class AthleteData:
    def __init__(self):
        self.name = None
        self.ageGroup = ""
        self.event = ""
        self.division = ""

        self.running = [ 0, 0, 0, 0, 0, 0, 0, 0 ]
        self.skierg = 0
        self.sledPush = 0
        self.sledPull = 0
        self.burpeeBroadJump = 0
        self.row = 0
        self.farmersCarry = 0
        self.sandbagLunges = 0
        self.wallBalls = 0
        self.roxzoneTime = 0
        self.runTotal = 0
        self.penalty = ""
        
        self.totalTime = 0

    def getCSVHeader():
        str = f'Name;Age Group;Event;Division;Running 1;1000m Ski Erg;Running 2;50m Sled Push;Running 3;50m Sled Pull;Running 4;80m Burpee Broad Jump;Running 5;1000m Row;Running 6;200m Farmers Carry;Running 7;100m Sandbag Lunges;Running 8;Wall Balls;Total Roxzone Time;Penalty;Total Time'
        return str

    def getCSVLine(self):
        #str = f'\"{self.name}\";\"{self.ageGroup}\";\"{self.event}\";\"{self.division}\";{timeSecondsToStr(self.running[0])};{timeSecondsToStr(self.skierg)};{timeSecondsToStr(self.running[1])};{timeSecondsToStr(self.sledPush)};{timeSecondsToStr(self.running[2])};{timeSecondsToStr(self.sledPull)};{timeSecondsToStr(self.running[3])};{timeSecondsToStr(self.burpeeBroadJump)};{timeSecondsToStr(self.running[4])};{timeSecondsToStr(self.row)};{timeSecondsToStr(self.running[5])};{timeSecondsToStr(self.farmersCarry)};{timeSecondsToStr(self.running[6])};{timeSecondsToStr(self.sandbagLunges)};{timeSecondsToStr(self.running[7])};{timeSecondsToStr(self.wallBalls)};{timeSecondsToStr(self.roxzoneTime)};{self.penalty};{timeSecondsToStr(self.totalTime)}'
        str = f'\"{self.name}\";\"{self.ageGroup}\";\"{self.event}\";\"{self.division}\";{self.running[0]};{self.skierg};{self.running[1]};{self.sledPush};{self.running[2]};{self.sledPull};{self.running[3]};{self.burpeeBroadJump};{self.running[4]};{self.row};{self.running[5]};{self.farmersCarry};{self.running[6]};{self.sandbagLunges};{self.running[7]};{self.wallBalls};{self.roxzoneTime};{self.penalty};{self.totalTime}'
        return str

    def writeHeaderInExcelWorksheet(sheet: Worksheet):
        # write the headers
        sheet.cell(1, 1).value = "Name"
        sheet.cell(1, 2).value = "Age Group"
        sheet.cell(1, 3).value = "Event"
        sheet.cell(1, 4).value = "Division"
        sheet.cell(1, 5).value = "Running 1"
        sheet.cell(1, 6).value = "1000m Ski Erg"
        sheet.cell(1, 7).value = "Running 2"
        sheet.cell(1, 8).value = "50m Sled Push"
        sheet.cell(1, 9).value = "Running 3"
        sheet.cell(1, 10).value = "50m Sled Pull"
        sheet.cell(1, 11).value = "Running 4"
        sheet.cell(1, 12).value = "80m Burpee Broad Jump"
        sheet.cell(1, 13).value = "Running 5"
        sheet.cell(1, 14).value = "1000m Row"
        sheet.cell(1, 15).value = "Running 6"
        sheet.cell(1, 16).value = "200m Farmer's Carry"
        sheet.cell(1, 17).value = "Running 7"
        sheet.cell(1, 18).value = "100m Sandbag Lunges"
        sheet.cell(1, 19).value = "Running 8"
        sheet.cell(1, 20).value = "Wall Balls"
        sheet.cell(1, 21).value = "Roxzone Time"
        sheet.cell(1, 22).value = "Penalty"
        sheet.cell(1, 23).value = "Total Time"

        # set all the cells as bold
        bold_font = Font(bold=True)
        row = sheet[1]
        for cell in row:
            cell.font = bold_font

    def writeAtheleteInfoInWorksheet(self, sheet: Worksheet, row: int):
        sheet.cell(row, 1).value = self.name
        sheet.cell(row, 2).value = self.ageGroup
        sheet.cell(row, 3).value = self.event
        sheet.cell(row, 4).value = self.division
        sheet.cell(row, 5).value = timeSecondsToDeltaTime(self.running[0])
        sheet.cell(row, 6).value = timeSecondsToDeltaTime(self.skierg)
        sheet.cell(row, 7).value = timeSecondsToDeltaTime(self.running[1])
        sheet.cell(row, 8).value = timeSecondsToDeltaTime(self.sledPush)
        sheet.cell(row, 9).value = timeSecondsToDeltaTime(self.running[2])
        sheet.cell(row, 10).value = timeSecondsToDeltaTime(self.sledPull)
        sheet.cell(row, 11).value = timeSecondsToDeltaTime(self.running[3])
        sheet.cell(row, 12).value = timeSecondsToDeltaTime(self.burpeeBroadJump)
        sheet.cell(row, 13).value = timeSecondsToDeltaTime(self.running[4])
        sheet.cell(row, 14).value = timeSecondsToDeltaTime(self.row)
        sheet.cell(row, 15).value = timeSecondsToDeltaTime(self.running[5])
        sheet.cell(row, 16).value = timeSecondsToDeltaTime(self.farmersCarry)
        sheet.cell(row, 17).value = timeSecondsToDeltaTime(self.running[6])
        sheet.cell(row, 18).value = timeSecondsToDeltaTime(self.sandbagLunges)
        sheet.cell(row, 19).value = timeSecondsToDeltaTime(self.running[7])
        sheet.cell(row, 20).value = timeSecondsToDeltaTime(self.wallBalls)
        sheet.cell(row, 21).value = timeSecondsToDeltaTime(self.roxzoneTime)
        sheet.cell(row, 22).value = self.penalty
        sheet.cell(row, 23).value = timeSecondsToDeltaTime(self.totalTime)


# Writes a file with the info of all the athletes in the given list
def writeAtheletesToCSV(athletes: list[AthleteData], filePath: str):
    f = open(file=filePath, mode="w", encoding="utf-8")
    f.write(AthleteData.getCSVHeader() + '\n')
    for athlete in athletes:
        f.write(athlete.getCSVLine() + '\n')
    f.flush()
    f.close()
    print(f'Finishing writing {len(athletes)} athletes to file {filePath}')


def fillExcelWorksheet(sheet: Worksheet, athletes:list[AthleteData]):
    # write the header first
    AthleteData.writeHeaderInExcelWorksheet(sheet)

    # now write all the athlete data
    row = 2
    for athlete in athletes:
        athlete.writeAtheleteInfoInWorksheet(sheet, row)
        row += 1
